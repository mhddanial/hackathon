"""
prepare_data.py - Batam SmartFlow Data Cleaning & Import Script
================================================================
Membaca data dari Excel, membersihkan, lalu mengupload ke Supabase.

Usage (dari direktori backend/):
    python scripts/prepare_data.py

Requirements:
    pip install openpyxl supabase python-dotenv
"""

import os
import sys
import openpyxl
from datetime import time as dt_time
from dotenv import load_dotenv
from supabase import create_client, Client

# Load env
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), '../.env'))

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_ANON_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERROR: SUPABASE_URL atau SUPABASE_ANON_KEY tidak ditemukan di .env")
    sys.exit(1)

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Path ke Excel
EXCEL_PATH = os.path.join(
    os.path.dirname(__file__),
    '../../docs/Batam SmartFlow Data Master (1).xlsx'
)

if not os.path.exists(EXCEL_PATH):
    print(f"ERROR: File Excel tidak ditemukan di: {EXCEL_PATH}")
    sys.exit(1)

print(f"[INFO] Membaca file: {os.path.abspath(EXCEL_PATH)}")
wb = openpyxl.load_workbook(EXCEL_PATH)


# ==============================================================================
# 1. Bersihkan & upload road_segments
# ==============================================================================
print("\n-- 1. road_segments --------------------------------------------------")

ws_seg = wb['road_segments']
segments_data = []

for row in ws_seg.iter_rows(min_row=2, values_only=True):
    segment_id, name, corridor_type, origin_name, destination_name, \
        start_lat, start_lng, end_lat, end_lng, length_km, typical_speed_kmh = row

    # Skip baris kosong
    if not segment_id:
        continue

    # Pastikan corridor_type lowercase dan tanpa spasi ekstra
    corridor_type_clean = str(corridor_type).strip().lower().replace(' ', '_')

    segments_data.append({
        "segment_id":        str(segment_id).strip(),
        "name":              str(name).strip(),
        "corridor_type":     corridor_type_clean,
        "origin_name":       str(origin_name).strip() if origin_name else None,
        "destination_name":  str(destination_name).strip() if destination_name else None,
        "start_lat":         float(start_lat),
        "start_lng":         float(start_lng),
        "end_lat":           float(end_lat),
        "end_lng":           float(end_lng),
        "length_km":         float(length_km),
        "typical_speed_kmh": float(typical_speed_kmh),
    })

print(f"   Ditemukan {len(segments_data)} segmen:")
for s in segments_data:
    print(f"   - {s['segment_id']}: {s['name']} [{s['corridor_type']}]")

# Upload ke Supabase (upsert by segment_id)
print("   >> Uploading ke Supabase...")
resp = supabase.table("road_segments").upsert(
    segments_data,
    on_conflict="segment_id"
).execute()
print(f"   [OK] road_segments: {len(resp.data)} baris berhasil di-upsert")

# Simpan mapping segment_id untuk validasi
valid_segment_ids = {s['segment_id'] for s in segments_data}

# ==============================================================================
# English description overrides
# Key: (segment_id, day_type, hour_start)
# ==============================================================================
DESCRIPTION_EN = {
    ('SEG001', 'weekday',  6): 'Morning peak: Container truck and factory worker congestion',
    ('SEG001', 'weekday',  9): 'Normal logistics operating hours',
    ('SEG001', 'weekday', 15): 'Evening peak & Batu Ampar cargo vessel cut-off window',
    ('SEG001', 'weekday', 19): 'Free-flowing traffic (off-peak night hours)',
    ('SEG001', 'weekend',  0): 'Weekend: Minimal factory activity, corridor running freely',
    ('SEG002', 'weekday',  6): 'Heavy private vehicle and public transport congestion',
    ('SEG002', 'weekday',  8): 'Moderate urban traffic flow',
    ('SEG002', 'weekday', 16): 'Evening rush hour congestion',
    ('SEG002', 'weekday', 18): 'Free-flowing traffic (night and early morning)',
    ('SEG002', 'weekend',  0): 'Weekend: Residents commuting to shopping centers',
    ('SEG003', 'weekday',  6): 'Batamindo Industrial Park inbound and outbound traffic',
    ('SEG003', 'weekday',  8): 'Moderate industrial zone logistics traffic',
    ('SEG003', 'weekday', 16): 'Factory shift change congestion',
    ('SEG003', 'weekday', 18): 'Quiet industrial zone traffic (off-peak)',
    ('SEG003', 'weekend',  0): 'Weekend: Minimal Batamindo operations',
    ('SEG004', 'weekday',  6): 'Normal shipyard logistics operations',
    ('SEG004', 'weekday', 15): 'Heavy transport buildup at Tanjung Uncang shipyard',
    ('SEG004', 'weekday', 18): 'Free-flowing Tanjung Uncang corridor',
    ('SEG004', 'weekend',  0): 'Weekend: Shipyard operations on holiday',
    ('SEG005', 'weekday',  0): 'Barelang access road free-flowing on weekdays',
    ('SEG005', 'weekend',  8): 'Weekend: Barelang tourist traffic',
    ('SEG005', 'weekend', 17): 'Weekend evening: free-flowing traffic',
    ('SEG006', 'weekday',  6): 'Normal Hang Nadim Airport traffic',
    ('SEG006', 'weekday', 18): 'Light airport traffic at night',
    ('SEG006', 'weekend',  0): 'Weekend: Free-flowing traffic to the airport',
}


# ==============================================================================
# 2. Bersihkan & upload congestion_multipliers
# ==============================================================================
print("\n-- 2. congestion_multipliers -----------------------------------------")

ws_mult = wb['congestion_multipliers']
multipliers_data = []
skipped = []
duplicates_check = set()

for row in ws_mult.iter_rows(min_row=2, min_col=1, max_col=8, values_only=True):
    segment_id, day_type, hour_window, hour_start, hour_end, \
        multiplier, congestion_level, description = row

    # Skip baris yang tidak punya segment_id (baris kosong / catatan)
    if not segment_id:
        continue

    segment_id = str(segment_id).strip()
    day_type = str(day_type).strip().lower()

    # Validasi segment_id ada di road_segments
    if segment_id not in valid_segment_ids:
        skipped.append(f"{segment_id} - segment_id tidak ditemukan di road_segments")
        continue

    # Validasi day_type
    if day_type not in ('weekday', 'weekend'):
        skipped.append(f"{segment_id} day_type='{day_type}' tidak valid (harus weekday/weekend)")
        continue

    # Cast hour_start / hour_end ke integer
    try:
        hour_start_int = int(float(hour_start))
        hour_end_int   = int(float(hour_end))
    except (TypeError, ValueError):
        skipped.append(f"{segment_id} {day_type} hour_start/end tidak valid: {hour_start}, {hour_end}")
        continue

    # Cek duplikat
    key = (segment_id, day_type, hour_start_int)
    if key in duplicates_check:
        skipped.append(f"DUPLIKAT: {segment_id} {day_type} hour_start={hour_start_int}")
        continue
    duplicates_check.add(key)

    multipliers_data.append({
        "segment_id":       segment_id,
        "day_type":         day_type,
        "hour_window":      str(hour_window).strip() if hour_window else None,
        "hour_start":       hour_start_int,
        "hour_end":         hour_end_int,
        "multiplier":       float(multiplier),
        "congestion_level": str(congestion_level).strip().upper(),
        "description":      DESCRIPTION_EN.get(
                                (segment_id, day_type, hour_start_int),
                                str(description).strip() if description else None
                            ),
    })

print(f"   Ditemukan {len(multipliers_data)} baris valid")

if skipped:
    print(f"   [WARN] {len(skipped)} baris di-skip:")
    for s in skipped:
        print(f"      - {s}")

# Coverage check sebelum upload
coverage = {}
for row in multipliers_data:
    sid = row['segment_id']
    dt  = row['day_type']
    if sid not in coverage:
        coverage[sid] = set()
    coverage[sid].add(dt)

print("\n   Coverage matrix:")
print(f"   {'segment_id':<10} {'weekday':>10} {'weekend':>10}")
for sid in sorted(valid_segment_ids):
    has_wd = '[OK]'   if 'weekday' in coverage.get(sid, set()) else '[MISS]'
    has_we = '[OK]'   if 'weekend' in coverage.get(sid, set()) else '[MISS]'
    print(f"   {sid:<10} {has_wd:>10} {has_we:>10}")

missing_coverage = [
    sid for sid in valid_segment_ids
    if 'weekday' not in coverage.get(sid, set()) or 'weekend' not in coverage.get(sid, set())
]
if missing_coverage:
    print(f"\n   [WARN] Segmen tanpa coverage weekday+weekend: {missing_coverage}")
    print("      Training akan dilanjutkan tapi model tidak bisa prediksi kombinasi ini.")
else:
    print("\n   [OK] Semua segmen punya data weekday & weekend")

# Upload ke Supabase
print("   >> Uploading ke Supabase...")
resp = supabase.table("congestion_multipliers").upsert(
    multipliers_data,
    on_conflict="segment_id,day_type,hour_start"
).execute()
print(f"   [OK] congestion_multipliers: {len(resp.data)} baris berhasil di-upsert")


# ==============================================================================
# 3. Bersihkan & upload ferry_schedules
# ==============================================================================
print("\n-- 3. ferry_schedules ------------------------------------------------")

ws_ferry = wb['ferry_schedules']
ferry_data = []


def time_to_str(t):
    """Konversi datetime.time ke string HH:MM:SS"""
    if t is None:
        return None
    if isinstance(t, dt_time):
        return t.strftime("%H:%M:%S")
    return str(t)


for row in ws_ferry.iter_rows(min_row=2, min_col=1, max_col=8, values_only=True):
    terminal_id, terminal, terminal_name, destination, vessel_type, \
        departure_time, cutoff_time, cargo_capacity_tons = row

    if not terminal_id:
        continue

    ferry_data.append({
        "terminal_id":          str(terminal_id).strip(),
        "terminal":             str(terminal).strip(),
        "terminal_name":        str(terminal_name).strip() if terminal_name else None,
        "destination":          str(destination).strip(),
        "vessel_type":          str(vessel_type).strip() if vessel_type else None,
        "departure_time":       time_to_str(departure_time),
        "cutoff_time":          time_to_str(cutoff_time),
        "cargo_capacity_tons":  float(cargo_capacity_tons) if cargo_capacity_tons is not None else None,
    })

print(f"   Ditemukan {len(ferry_data)} jadwal ferry")
for f in ferry_data:
    print(f"   - [{f['terminal_id']}] {f['terminal']} dep={f['departure_time']}"
          f" cutoff={f['cutoff_time']} cap={f['cargo_capacity_tons']}t [{f['vessel_type']}]")

print("   >> Uploading ke Supabase...")
resp = supabase.table("ferry_schedules").upsert(ferry_data).execute()
print(f"   [OK] ferry_schedules: {len(resp.data)} baris berhasil di-upsert")


# ==============================================================================
# Ringkasan Final
# ==============================================================================
print("\n" + "="*60)
print("[DONE] IMPORT SELESAI - Ringkasan")
print("="*60)
print(f"   road_segments         : {len(segments_data)} baris")
print(f"   congestion_multipliers: {len(multipliers_data)} baris")
print(f"   ferry_schedules       : {len(ferry_data)} baris")
print()
print("Langkah selanjutnya:")
print("   1. Buka backend/scripts/train_model.ipynb")
print("   2. Jalankan sel 1 (Fetch Data) -> cek row count sesuai di atas")
print("   3. Jalankan sel 2 (Sanity Check Coverage) -> semua segmen harus terisi")
print("   4. Lanjut sel 3-9 (Generate Synthetic -> Train -> Export)")
print("="*60)
